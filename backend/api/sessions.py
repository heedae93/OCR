"""
Session management API endpoints
"""
import logging
import json
import zipfile

# fcntl은 Unix 전용 - Windows에서는 파일 잠금 없이 동작
try:
    import fcntl
    _HAS_FCNTL = True
except ImportError:
    _HAS_FCNTL = False
import io
import os
from typing import List, Optional
from datetime import datetime
from pathlib import Path
import shutil

from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session as DBSession
from PyPDF2 import PdfMerger, PdfReader, PdfWriter

from database import get_db, Session, SessionDocument, Job, SessionLocal, User
from utils.file_utils import generate_unique_id
from config import Config

# Export status directory
EXPORT_STATUS_DIR = Config.TEMP_DIR / "export_status"
EXPORT_STATUS_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger(__name__)

router = APIRouter()


# Pydantic models for request/response
class SessionCreate(BaseModel):
    """Request model for creating a session"""
    session_name: str
    description: Optional[str] = None


class SessionUpdate(BaseModel):
    """Request model for updating a session"""
    session_name: Optional[str] = None
    description: Optional[str] = None


class DocumentInSession(BaseModel):
    """Document info within a session"""
    job_id: str
    original_filename: str
    status: str
    progress_percent: float
    current_page: int
    total_pages: int
    order: int
    is_selected: bool
    pdf_url: Optional[str] = None
    added_at: str


class SessionResponse(BaseModel):
    """Session response model"""
    session_id: str
    session_name: str
    description: Optional[str]
    created_at: str
    updated_at: str
    total_documents: int
    completed_documents: int
    documents: List[DocumentInSession]


class AddDocumentRequest(BaseModel):
    """Request to add a document to session"""
    job_id: str


class UpdateSelectionRequest(BaseModel):
    """Request to update document selection state"""
    job_ids: List[str]
    is_selected: bool


class JobOrderItem(BaseModel):
    """Single job order item"""
    job_id: str
    order: int


class ReorderDocumentsRequest(BaseModel):
    """Request to reorder documents in a session"""
    job_orders: List[JobOrderItem]


class ExportStatusResponse(BaseModel):
    """Export job status response"""
    export_id: str
    status: str  # pending, processing, completed, failed
    progress_percent: float
    current_file: int
    total_files: int
    current_page: Optional[int] = None
    total_pages: Optional[int] = None
    message: Optional[str] = None
    download_url: Optional[str] = None


class MultiFormatExportRequest(BaseModel):
    """Request for multi-format session export"""
    formats: List[str]  # ['pdf', 'txt', 'xml', 'json']
    as_zip: bool = True


def save_export_status(export_id: str, status: dict):
    """Save export status to file"""
    status_file = EXPORT_STATUS_DIR / f"{export_id}.json"
    status["updated_at"] = datetime.now().isoformat()
    try:
        with open(status_file, 'w') as f:
            if _HAS_FCNTL:
                fcntl.flock(f.fileno(), fcntl.LOCK_EX)
            json.dump(status, f)
            f.flush()  # Flush buffer to OS
            os.fsync(f.fileno())  # Ensure data is written to disk
            if _HAS_FCNTL:
                fcntl.flock(f.fileno(), fcntl.LOCK_UN)
        logger.debug(f"Export status saved: {export_id} -> {status.get('status')}")
    except Exception as e:
        logger.error(f"Failed to save export status: {e}")


def read_export_status(export_id: str) -> Optional[dict]:
    """Read export status from file"""
    status_file = EXPORT_STATUS_DIR / f"{export_id}.json"
    if not status_file.exists():
        return None
    try:
        with open(status_file, 'r') as f:
            if _HAS_FCNTL:
                fcntl.flock(f.fileno(), fcntl.LOCK_SH)
            data = json.load(f)
            if _HAS_FCNTL:
                fcntl.flock(f.fileno(), fcntl.LOCK_UN)
            return data
    except Exception as e:
        logger.error(f"Failed to read export status: {e}")
        return None


@router.post("/sessions", response_model=SessionResponse)
async def create_session(
    request: SessionCreate,
    user_id: str = Config.DEFAULT_USER_ID,
    db: DBSession = Depends(get_db)
):
    """Create a new session"""
    try:
        session_id = generate_unique_id()

        # Find the user where 'username' or 'user_id' is 'futurenuri'
        futurenuri_user = db.query(User).filter(
            (User.username == "futurenuri") | (User.user_id == "futurenuri")
        ).first()

        actual_user_id = futurenuri_user.user_id if futurenuri_user else user_id

        new_session = Session(
            session_id=session_id,
            user_id=actual_user_id,
            session_name=request.session_name,
            description=request.description
        )

        db.add(new_session)
        db.commit()
        db.refresh(new_session)

        logger.info(f"Created session: {session_id} - {request.session_name}")

        return SessionResponse(
            session_id=new_session.session_id,
            session_name=new_session.session_name,
            description=new_session.description,
            created_at=new_session.created_at.isoformat(),
            updated_at=new_session.updated_at.isoformat(),
            total_documents=0,
            completed_documents=0,
            documents=[]
        )

    except Exception as e:
        logger.error(f"Failed to create session: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions", response_model=List[SessionResponse])
async def list_sessions(
    user_id: str = Config.DEFAULT_USER_ID,
    db: DBSession = Depends(get_db)
):
    """List all sessions for a user"""
    try:
        sessions = db.query(Session).filter_by(user_id=user_id).order_by(Session.updated_at.desc()).all()

        result = []
        for session in sessions:
            # Get documents in this session
            session_docs = db.query(SessionDocument, Job).join(
                Job, SessionDocument.job_id == Job.job_id
            ).filter(
                SessionDocument.session_id == session.session_id
            ).order_by(SessionDocument.order).all()

            documents = []
            completed_count = 0

            for sess_doc, job in session_docs:
                # For processing/queued jobs, read latest status from file (non-blocking)
                status = job.status
                progress_percent = job.progress_percent or 0
                current_page = job.current_page or 0
                total_pages = job.total_pages or 0

                # Check file status for any non-completed job (processing, queued, etc.)
                if job.status in ("processing", "queued", "pending"):
                    from utils.job_manager import JobManager
                    file_status = JobManager.read_status_from_file(job.job_id)
                    if file_status:
                        status = file_status.get("status", job.status)
                        progress_percent = file_status.get("progress_percent", progress_percent)
                        current_page = file_status.get("current_page", current_page)
                        total_pages = file_status.get("total_pages", total_pages)

                if status == "completed":
                    completed_count += 1

                pdf_url = None
                if job.pdf_file_path:
                    import os
                    filename = os.path.basename(job.pdf_file_path)
                    pdf_url = f"/files/processed/{filename}"

                documents.append(DocumentInSession(
                    job_id=job.job_id,
                    original_filename=job.original_filename,
                    status=status,
                    progress_percent=progress_percent,
                    current_page=current_page,
                    total_pages=total_pages,
                    order=sess_doc.order,
                    is_selected=sess_doc.is_selected,
                    pdf_url=pdf_url,
                    added_at=sess_doc.added_at.isoformat()
                ))

            result.append(SessionResponse(
                session_id=session.session_id,
                session_name=session.session_name,
                description=session.description,
                created_at=session.created_at.isoformat(),
                updated_at=session.updated_at.isoformat(),
                total_documents=len(documents),
                completed_documents=completed_count,
                documents=documents
            ))

        return result

    except Exception as e:
        logger.error(f"Failed to list sessions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{session_id}", response_model=SessionResponse)
async def get_session(
    session_id: str,
    db: DBSession = Depends(get_db)
):
    """Get a specific session with all its documents"""
    try:
        session = db.query(Session).filter_by(session_id=session_id).first()

        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Get documents in this session
        session_docs = db.query(SessionDocument, Job).join(
            Job, SessionDocument.job_id == Job.job_id
        ).filter(
            SessionDocument.session_id == session_id
        ).order_by(SessionDocument.order).all()

        documents = []
        completed_count = 0

        for sess_doc, job in session_docs:
            if job.status == "completed":
                completed_count += 1

            pdf_url = None
            if job.pdf_file_path:
                import os
                filename = os.path.basename(job.pdf_file_path)
                pdf_url = f"/files/processed/{filename}"

            documents.append(DocumentInSession(
                job_id=job.job_id,
                original_filename=job.original_filename,
                status=job.status,
                progress_percent=job.progress_percent,
                current_page=job.current_page,
                total_pages=job.total_pages,
                order=sess_doc.order,
                is_selected=sess_doc.is_selected,
                pdf_url=pdf_url,
                added_at=sess_doc.added_at.isoformat()
            ))

        return SessionResponse(
            session_id=session.session_id,
            session_name=session.session_name,
            description=session.description,
            created_at=session.created_at.isoformat(),
            updated_at=session.updated_at.isoformat(),
            total_documents=len(documents),
            completed_documents=completed_count,
            documents=documents
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get session: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sessions/{session_id}", response_model=SessionResponse)
async def update_session(
    session_id: str,
    request: SessionUpdate,
    db: DBSession = Depends(get_db)
):
    """Update session metadata"""
    try:
        session = db.query(Session).filter_by(session_id=session_id).first()

        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        if request.session_name is not None:
            session.session_name = request.session_name
        if request.description is not None:
            session.description = request.description

        session.updated_at = datetime.now()

        db.commit()
        db.refresh(session)

        # Return updated session
        return await get_session(session_id, db)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update session: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sessions/{session_id}")
async def delete_session(
    session_id: str,
    db: DBSession = Depends(get_db)
):
    """Delete a session (does not delete the jobs, only the session grouping)"""
    try:
        session = db.query(Session).filter_by(session_id=session_id).first()

        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        db.delete(session)
        db.commit()

        logger.info(f"Deleted session: {session_id}")

        return {"message": "Session deleted successfully", "session_id": session_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete session: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sessions/{session_id}/documents")
async def add_document_to_session(
    session_id: str,
    request: AddDocumentRequest,
    db: DBSession = Depends(get_db)
):
    """Add a document (job) to a session"""
    try:
        # Check if session exists
        session = db.query(Session).filter_by(session_id=session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Check if job exists
        job = db.query(Job).filter_by(job_id=request.job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        # Check if already in session
        existing = db.query(SessionDocument).filter_by(
            session_id=session_id,
            job_id=request.job_id
        ).first()

        if existing:
            raise HTTPException(status_code=400, detail="Document already in session")

        # Get max order
        max_order = db.query(SessionDocument).filter_by(
            session_id=session_id
        ).count()

        # Add to session
        session_doc = SessionDocument(
            session_id=session_id,
            job_id=request.job_id,
            order=max_order,
            is_selected=True
        )

        db.add(session_doc)
        session.updated_at = datetime.now()
        db.commit()

        logger.info(f"Added document {request.job_id} to session {session_id}")

        return {"message": "Document added to session", "job_id": request.job_id, "session_id": session_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to add document to session: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/sessions/{session_id}/documents/{job_id}")
async def remove_document_from_session(
    session_id: str,
    job_id: str,
    db: DBSession = Depends(get_db)
):
    """Remove a document from a session"""
    try:
        session_doc = db.query(SessionDocument).filter_by(
            session_id=session_id,
            job_id=job_id
        ).first()

        if not session_doc:
            raise HTTPException(status_code=404, detail="Document not in session")

        db.delete(session_doc)

        # Update session timestamp
        session = db.query(Session).filter_by(session_id=session_id).first()
        if session:
            session.updated_at = datetime.now()

        db.commit()

        logger.info(f"Removed document {job_id} from session {session_id}")

        return {"message": "Document removed from session", "job_id": job_id, "session_id": session_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to remove document from session: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sessions/{session_id}/selection")
async def update_document_selection(
    session_id: str,
    request: UpdateSelectionRequest,
    db: DBSession = Depends(get_db)
):
    """Update selection state for multiple documents in a session"""
    try:
        # Update selection state for all specified documents
        updated_count = 0
        for job_id in request.job_ids:
            session_doc = db.query(SessionDocument).filter_by(
                session_id=session_id,
                job_id=job_id
            ).first()

            if session_doc:
                session_doc.is_selected = request.is_selected
                updated_count += 1

        # Update session timestamp
        session = db.query(Session).filter_by(session_id=session_id).first()
        if session:
            session.updated_at = datetime.now()

        db.commit()

        logger.info(f"Updated selection for {updated_count} documents in session {session_id}")

        return {
            "message": "Selection updated",
            "session_id": session_id,
            "updated_count": updated_count,
            "is_selected": request.is_selected
        }

    except Exception as e:
        logger.error(f"Failed to update selection: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


def do_export_merge(export_id: str, session_id: str, session_name: str, pdf_paths: List[str]):
    """Background task to merge PDFs with page-level progress tracking"""
    total_files = len(pdf_paths)

    try:
        # Initialize status
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "processing",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": total_files,
            "message": "Scanning PDFs..."
        })

        # First pass: count total pages
        pdf_info = []  # [(path, page_count), ...]
        total_pages = 0
        for idx, pdf_path in enumerate(pdf_paths):
            try:
                reader = PdfReader(str(pdf_path))
                page_count = len(reader.pages)
                pdf_info.append((pdf_path, reader, page_count))
                total_pages += page_count
                save_export_status(export_id, {
                    "export_id": export_id,
                    "session_id": session_id,
                    "status": "processing",
                    "progress_percent": (idx + 1) / total_files * 5,  # 0-5% for scanning
                    "current_file": idx + 1,
                    "total_files": total_files,
                    "message": f"Scanning file {idx + 1}/{total_files} ({page_count} pages)"
                })
            except Exception as e:
                logger.warning(f"Export {export_id}: Failed to read {pdf_path}: {e}")

        logger.info(f"Export {export_id}: Total {total_pages} pages from {len(pdf_info)} files")

        # Second pass: merge pages with progress
        writer = PdfWriter()
        pages_processed = 0

        for file_idx, (pdf_path, reader, page_count) in enumerate(pdf_info):
            for page_idx in range(page_count):
                try:
                    writer.add_page(reader.pages[page_idx])
                    pages_processed += 1

                    # Update progress every 10 pages or at end of file
                    if pages_processed % 10 == 0 or page_idx == page_count - 1:
                        progress = 5 + (pages_processed / total_pages) * 90  # 5-95% for merging
                        save_export_status(export_id, {
                            "export_id": export_id,
                            "session_id": session_id,
                            "status": "processing",
                            "progress_percent": progress,
                            "current_file": file_idx + 1,
                            "total_files": total_files,
                            "current_page": pages_processed,
                            "total_pages": total_pages,
                            "message": f"Merging page {pages_processed}/{total_pages}"
                        })
                except Exception as e:
                    logger.warning(f"Export {export_id}: Failed to add page {page_idx} from {pdf_path}: {e}")

            logger.info(f"Export {export_id}: Merged {pdf_path} ({page_count} pages)")

        # Write to file
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "processing",
            "progress_percent": 95,
            "current_file": total_files,
            "total_files": total_files,
            "current_page": total_pages,
            "total_pages": total_pages,
            "message": f"Writing {total_pages} pages to PDF..."
        })

        output_filename = f"{export_id}_merged.pdf"
        output_path = Config.PROCESSED_DIR / output_filename

        with open(output_path, 'wb') as output_file:
            writer.write(output_file)

        logger.info(f"Export {export_id}: Created merged PDF: {output_path}")

        # Complete
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "completed",
            "progress_percent": 100,
            "current_file": total_files,
            "total_files": total_files,
            "current_page": total_pages,
            "total_pages": total_pages,
            "message": "Export completed",
            "download_url": f"/sessions/export-download/{export_id}",
            "filename": f"{session_name}_merged.pdf"
        })

    except Exception as e:
        logger.error(f"Export {export_id} failed: {e}")
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "failed",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": total_files,
            "message": str(e)
        })


@router.post("/sessions/{session_id}/export-merged")
async def export_session_merged(
    session_id: str,
    background_tasks: BackgroundTasks,
    db: DBSession = Depends(get_db)
):
    """
    Start async export of selected PDFs as a merged PDF.
    Returns export_id for progress tracking.
    """
    try:
        # Get session
        session = db.query(Session).filter_by(session_id=session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Get selected documents in order
        selected_docs = db.query(SessionDocument, Job).join(
            Job, SessionDocument.job_id == Job.job_id
        ).filter(
            SessionDocument.session_id == session_id,
            SessionDocument.is_selected == True,
            Job.status == "completed"
        ).order_by(SessionDocument.order).all()

        if not selected_docs:
            raise HTTPException(status_code=400, detail="No completed documents selected")

        # Collect PDF paths
        pdf_paths = []
        for sess_doc, job in selected_docs:
            pdf_path = None
            if job.final_pdf_path and Path(job.final_pdf_path).exists():
                pdf_path = job.final_pdf_path
            elif job.pdf_file_path and Path(job.pdf_file_path).exists():
                pdf_path = job.pdf_file_path

            if pdf_path:
                pdf_paths.append(pdf_path)
            else:
                logger.warning(f"PDF not found for job {job.job_id}")

        if not pdf_paths:
            raise HTTPException(status_code=404, detail="No PDF files found for selected documents")

        # Create export job
        export_id = generate_unique_id()

        # Initialize status
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "pending",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": len(pdf_paths),
            "message": "Export queued"
        })

        # Start background task
        background_tasks.add_task(do_export_merge, export_id, session_id, session.session_name, pdf_paths)

        logger.info(f"Started export {export_id} for session {session_id} with {len(pdf_paths)} files")

        return {
            "export_id": export_id,
            "status": "pending",
            "total_files": len(pdf_paths),
            "message": "Export started"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to start export: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/export-status/{export_id}", response_model=ExportStatusResponse)
async def get_export_status(export_id: str):
    """Get export job status"""
    status = read_export_status(export_id)
    if not status:
        raise HTTPException(status_code=404, detail="Export job not found")

    return ExportStatusResponse(
        export_id=status.get("export_id", export_id),
        status=status.get("status", "unknown"),
        progress_percent=status.get("progress_percent", 0),
        current_file=status.get("current_file", 0),
        total_files=status.get("total_files", 0),
        current_page=status.get("current_page"),
        total_pages=status.get("total_pages"),
        message=status.get("message"),
        download_url=status.get("download_url")
    )


@router.get("/sessions/export-download/{export_id}")
async def download_export(export_id: str):
    """Download completed export (PDF or ZIP)"""
    status = read_export_status(export_id)
    if not status:
        raise HTTPException(status_code=404, detail="Export job not found")

    if status.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Export not completed yet")

    # Try ZIP first (multi-format export), then PDF (merged export)
    output_path = Config.PROCESSED_DIR / f"{export_id}_export.zip"
    if output_path.exists():
        download_name = status.get("filename", f"{export_id}_export.zip")
        media_type = "application/zip"
    else:
        output_path = Config.PROCESSED_DIR / f"{export_id}_merged.pdf"
        download_name = status.get("filename", f"{export_id}_merged.pdf")
        media_type = "application/pdf"

    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Export file not found")

    # Get file size for Content-Length header
    file_size = output_path.stat().st_size
    logger.info(f"Serving export file: {output_path} ({file_size / 1024 / 1024:.1f}MB)")

    return FileResponse(
        path=str(output_path),
        filename=download_name,
        headers={"Content-Length": str(file_size)},
        media_type=media_type
    )


@router.get("/sessions/{session_id}/export-individual/{job_id}")
async def export_individual_pdf(
    session_id: str,
    job_id: str,
    db: DBSession = Depends(get_db)
):
    """
    Export a single PDF from a session
    """
    try:
        # Verify job is in session
        session_doc = db.query(SessionDocument).filter_by(
            session_id=session_id,
            job_id=job_id
        ).first()

        if not session_doc:
            raise HTTPException(status_code=404, detail="Document not in session")

        # Get job
        job = db.query(Job).filter_by(job_id=job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        if job.status != "completed":
            raise HTTPException(status_code=400, detail="Job not completed")

        # Get PDF path
        pdf_path = None
        if job.final_pdf_path and Path(job.final_pdf_path).exists():
            pdf_path = job.final_pdf_path
        elif job.pdf_file_path and Path(job.pdf_file_path).exists():
            pdf_path = job.pdf_file_path

        if not pdf_path:
            raise HTTPException(status_code=404, detail="PDF file not found")

        # Return file
        return FileResponse(
            path=str(pdf_path),
            filename=job.original_filename.replace(Path(job.original_filename).suffix, "_ocr.pdf"),
            media_type="application/pdf"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export individual PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/stats/overview")
async def get_stats_overview(
    user_id: str = Config.DEFAULT_USER_ID,
    db: DBSession = Depends(get_db)
):
    """
    Get overview statistics for dashboard
    """
    try:
        from sqlalchemy import func

        # Get all sessions
        sessions = db.query(Session).filter_by(user_id=user_id).all()

        # Get all jobs
        all_jobs = db.query(Job).filter_by(user_id=user_id).all()

        # Calculate stats
        total_sessions = len(sessions)
        total_documents = len(all_jobs)
        completed_documents = sum(1 for j in all_jobs if j.status == "completed")
        processing_documents = sum(1 for j in all_jobs if j.status == "processing")
        failed_documents = sum(1 for j in all_jobs if j.status == "failed")
        total_pages = sum(j.total_pages or 0 for j in all_jobs)
        total_text_blocks = sum(j.total_text_blocks or 0 for j in all_jobs)

        # Calculate average processing time
        completed_with_time = [j for j in all_jobs if j.status == "completed" and j.processing_time_seconds]
        avg_processing_time = (
            sum(j.processing_time_seconds for j in completed_with_time) / len(completed_with_time)
            if completed_with_time else 0
        )

        # Recent activity (last 7 days)
        from datetime import datetime, timedelta
        week_ago = datetime.now() - timedelta(days=7)
        recent_jobs = [j for j in all_jobs if j.created_at and j.created_at >= week_ago]

        return {
            "total_sessions": total_sessions,
            "total_documents": total_documents,
            "completed_documents": completed_documents,
            "processing_documents": processing_documents,
            "failed_documents": failed_documents,
            "total_pages": total_pages,
            "total_text_blocks": total_text_blocks,
            "success_rate": (completed_documents / total_documents * 100) if total_documents > 0 else 0,
            "average_processing_time": avg_processing_time,
            "recent_activity_count": len(recent_jobs)
        }

    except Exception as e:
        logger.error(f"Failed to get stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sessions/{session_id}/reorder")
async def reorder_documents(
    session_id: str,
    request: ReorderDocumentsRequest,
    db: DBSession = Depends(get_db)
):
    """Reorder documents within a session"""
    try:
        # Check if session exists
        session = db.query(Session).filter_by(session_id=session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Update order for each document
        updated_count = 0
        for job_order in request.job_orders:
            session_doc = db.query(SessionDocument).filter_by(
                session_id=session_id,
                job_id=job_order.job_id
            ).first()

            if session_doc:
                session_doc.order = job_order.order
                updated_count += 1

        # Update session timestamp
        session.updated_at = datetime.now()
        db.commit()

        logger.info(f"Reordered {updated_count} documents in session {session_id}")

        return {
            "message": "Documents reordered",
            "session_id": session_id,
            "updated_count": updated_count
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to reorder documents: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# Multi-format Session Export
# ============================================================

def get_ocr_json(job_id: str) -> Optional[dict]:
    """Load OCR results from JSON file"""
    json_path = Config.PROCESSED_DIR / f"{job_id}_ocr.json"
    if not json_path.exists():
        return None
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def generate_txt_content(ocr_data: dict) -> str:
    """Generate TXT content from OCR data"""
    lines = []
    for page in ocr_data.get('pages', []):
        page_num = page.get('page_number', 0)
        lines.append(f"\n{'='*60}")
        lines.append(f"[Page {page_num}]")
        lines.append(f"{'='*60}\n")

        page_lines = page.get('lines', [])
        if page_lines and page_lines[0].get('reading_order') is not None:
            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))

        for line in page_lines:
            text = line.get('text', '').strip()
            if text:
                lines.append(text)

    return '\n'.join(lines)


def generate_xml_content(ocr_data: dict, job_id: str) -> str:
    """Generate simple XML content from OCR data"""
    from xml.etree.ElementTree import Element, SubElement, tostring
    from xml.dom import minidom

    doc = Element('document')
    doc.set('job_id', job_id)
    doc.set('producer', 'BabelBrain OCR')

    for page in ocr_data.get('pages', []):
        page_elem = SubElement(doc, 'page')
        page_elem.set('number', str(page.get('page_number', 0)))
        page_elem.set('width', str(page.get('width', 0)))
        page_elem.set('height', str(page.get('height', 0)))

        page_lines = page.get('lines', [])
        if page_lines and page_lines[0].get('reading_order') is not None:
            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))

        for line in page_lines:
            line_elem = SubElement(page_elem, 'line')
            line_elem.set('confidence', str(line.get('confidence', 0)))

            bbox = line.get('bounding_box', {})
            if bbox:
                line_elem.set('x', str(bbox.get('x', 0)))
                line_elem.set('y', str(bbox.get('y', 0)))
                line_elem.set('width', str(bbox.get('width', 0)))
                line_elem.set('height', str(bbox.get('height', 0)))

            line_elem.text = line.get('text', '')

    xml_str = tostring(doc, encoding='unicode')
    return minidom.parseString(xml_str).toprettyxml(indent="  ")


def generate_excel_content(ocr_data: dict, job_id: str) -> Optional[bytes]:
    """Generate Excel statistics content from OCR data"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR Statistics"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Headers
        headers = ["Page", "Lines", "Characters", "Avg Confidence", "Min Conf", "Max Conf"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, page in enumerate(ocr_data.get('pages', []), 2):
            lines = page.get('lines', [])
            confs = [l.get('confidence', 0.95) for l in lines if l.get('confidence') is not None]
            confs = [c * 100 if c <= 1 else c for c in confs]
            char_count = sum(len(l.get('text', '')) for l in lines)

            ws.cell(row=row_idx, column=1, value=page.get('page_number', row_idx - 1))
            ws.cell(row=row_idx, column=2, value=len(lines))
            ws.cell(row=row_idx, column=3, value=char_count)
            ws.cell(row=row_idx, column=4, value=f"{sum(confs)/len(confs):.2f}%" if confs else "N/A")
            ws.cell(row=row_idx, column=5, value=f"{min(confs):.2f}%" if confs else "N/A")
            ws.cell(row=row_idx, column=6, value=f"{max(confs):.2f}%" if confs else "N/A")

        # Column widths
        for col_idx in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except ImportError:
        logger.warning("openpyxl not installed, skipping Excel export")
        return None
    except Exception as e:
        logger.warning(f"Failed to generate Excel content: {e}")
        return None


def do_multi_format_export(
    export_id: str,
    session_id: str,
    session_name: str,
    jobs: List[tuple],  # [(job_id, original_filename, pdf_path), ...]
    formats: List[str],
    as_zip: bool
):
    """Background task for multi-format export with progress tracking.

    Streams files directly to ZIP to avoid memory issues with large exports.
    """
    total_jobs = len(jobs)
    total_formats = len(formats)
    total_operations = total_jobs * total_formats

    try:
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "processing",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": total_jobs,
            "message": "Preparing export..."
        })

        # Create ZIP file and stream files directly to avoid memory issues
        output_path = Config.PROCESSED_DIR / f"{export_id}_export.zip"
        file_count = 0
        used_names = set()  # Track used filenames to handle duplicates
        operation_count = 0

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for job_idx, (job_id, original_filename, pdf_path) in enumerate(jobs):
                base_name = Path(original_filename).stem

                for fmt in formats:
                    operation_count += 1
                    progress = (operation_count / total_operations) * 95  # Leave 5% for finalization

                    save_export_status(export_id, {
                        "export_id": export_id,
                        "session_id": session_id,
                        "status": "processing",
                        "progress_percent": progress,
                        "current_file": job_idx + 1,
                        "total_files": total_jobs,
                        "message": f"Adding {fmt.upper()}: {base_name[:30]}..."
                    })

                    try:
                        # Generate unique filename to avoid duplicates
                        filename = f"{base_name}.{fmt}"
                        if filename in used_names:
                            # Append job_id prefix to make unique
                            filename = f"{job_id[:8]}_{base_name}.{fmt}"
                        used_names.add(filename)

                        if fmt == 'pdf':
                            if pdf_path and Path(pdf_path).exists():
                                # Use write() instead of writestr() for file paths (streams from disk)
                                zf.write(pdf_path, filename)
                                file_count += 1

                        elif fmt == 'txt':
                            ocr_data = get_ocr_json(job_id)
                            if ocr_data:
                                content = generate_txt_content(ocr_data)
                                zf.writestr(filename, content.encode('utf-8'))
                                file_count += 1

                        elif fmt == 'xml':
                            ocr_data = get_ocr_json(job_id)
                            if ocr_data:
                                content = generate_xml_content(ocr_data, job_id)
                                zf.writestr(filename, content.encode('utf-8'))
                                file_count += 1

                        elif fmt == 'json':
                            json_path = Config.PROCESSED_DIR / f"{job_id}_ocr.json"
                            if json_path.exists():
                                zf.write(json_path, filename)
                                file_count += 1

                        elif fmt == 'excel':
                            ocr_data = get_ocr_json(job_id)
                            if ocr_data:
                                content = generate_excel_content(ocr_data, job_id)
                                if content:
                                    excel_filename = f"{base_name}_statistics.xlsx"
                                    if excel_filename in used_names:
                                        excel_filename = f"{job_id[:8]}_{base_name}_statistics.xlsx"
                                    used_names.add(excel_filename)
                                    zf.writestr(excel_filename, content)
                                    file_count += 1

                    except Exception as e:
                        logger.warning(f"Failed to add {fmt} for {job_id}: {e}")

        if file_count == 0:
            # Remove empty ZIP
            if output_path.exists():
                output_path.unlink()
            save_export_status(export_id, {
                "export_id": export_id,
                "session_id": session_id,
                "status": "failed",
                "progress_percent": 0,
                "current_file": 0,
                "total_files": total_jobs,
                "message": "No files to export"
            })
            return

        download_filename = f"{session_name}_export.zip"
        file_size = output_path.stat().st_size if output_path.exists() else 0
        logger.info(f"Export {export_id}: Created {output_path} ({file_size / 1024 / 1024:.1f}MB) with {file_count} files")

        logger.info(f"Export {export_id}: Saving completed status...")
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "completed",
            "progress_percent": 100,
            "current_file": total_jobs,
            "total_files": total_jobs,
            "message": "Export completed",
            "download_url": f"/sessions/export-download/{export_id}",
            "filename": download_filename,
            "file_count": file_count
        })
        logger.info(f"Export {export_id}: Export fully completed!")

    except Exception as e:
        import traceback
        logger.error(f"Multi-format export {export_id} failed: {e}\n{traceback.format_exc()}")
        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "failed",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": total_jobs,
            "message": str(e)
        })


@router.post("/sessions/{session_id}/export-multi")
async def export_session_multi_format(
    session_id: str,
    request: MultiFormatExportRequest,
    background_tasks: BackgroundTasks,
    db: DBSession = Depends(get_db)
):
    """
    Export selected documents in multiple formats.
    Returns export_id for progress tracking.
    """
    try:
        # Validate formats
        valid_formats = {'pdf', 'txt', 'xml', 'json', 'excel'}
        formats = [f.lower() for f in request.formats if f.lower() in valid_formats]
        if not formats:
            raise HTTPException(status_code=400, detail="No valid formats specified")

        # Get session
        session = db.query(Session).filter_by(session_id=session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")

        # Get selected completed documents
        selected_docs = db.query(SessionDocument, Job).join(
            Job, SessionDocument.job_id == Job.job_id
        ).filter(
            SessionDocument.session_id == session_id,
            SessionDocument.is_selected == True,
            Job.status == "completed"
        ).order_by(SessionDocument.order).all()

        if not selected_docs:
            raise HTTPException(status_code=400, detail="No completed documents selected")

        # Collect job info
        jobs = []
        for sess_doc, job in selected_docs:
            pdf_path = job.final_pdf_path or job.pdf_file_path
            jobs.append((job.job_id, job.original_filename, pdf_path))

        # Create export job
        export_id = generate_unique_id()

        save_export_status(export_id, {
            "export_id": export_id,
            "session_id": session_id,
            "status": "pending",
            "progress_percent": 0,
            "current_file": 0,
            "total_files": len(jobs),
            "message": "Export queued"
        })

        # Start background task
        background_tasks.add_task(
            do_multi_format_export,
            export_id,
            session_id,
            session.session_name,
            jobs,
            formats,
            request.as_zip
        )

        logger.info(f"Started multi-format export {export_id} for session {session_id}: {formats}")

        return {
            "export_id": export_id,
            "status": "pending",
            "total_files": len(jobs),
            "formats": formats,
            "message": "Export started"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to start multi-format export: {e}")
        raise HTTPException(status_code=500, detail=str(e))
