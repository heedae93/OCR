"""
Export API for TXT, XML, Excel outputs
ABBYY-compatible XML format
"""
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
import uuid

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel

from config import Config
from utils.job_manager import JobManager
from core.ctc_char_confidence import HeuristicCharConfidenceEstimator
from database import SessionLocal, DownloadHistory, FileVersion, Job as DBJob

logger = logging.getLogger(__name__)

# 문자별 confidence 추정기 (싱글톤)
char_confidence_estimator = HeuristicCharConfidenceEstimator()
router = APIRouter(prefix="/api/export", tags=["export"])

job_manager = JobManager()


def record_download(job_id: str, user_id: str, file_type: str, ip: Optional[str] = None):
    """다운로드 이력 기록 + 버전 자동 생성"""
    try:
        db = SessionLocal()

        # 다운로드 이력
        record = DownloadHistory(job_id=job_id, user_id=user_id, file_type=file_type, ip_address=ip)
        db.add(record)

        # 버전 자동 생성
        last = db.query(FileVersion).filter_by(job_id=job_id).order_by(FileVersion.version_number.desc()).first()
        next_num = (last.version_number + 1) if last else 1
        db_job = db.query(DBJob).filter_by(job_id=job_id).first()
        version = FileVersion(
            job_id=job_id,
            user_id=user_id,
            version_number=next_num,
            version_label=f"v{next_num}.0",
            note=f"{file_type.upper()} 내보내기 시 자동 생성",
            file_size_bytes=db_job.file_size_bytes if db_job else None,
        )
        db.add(version)

        db.commit()
        db.close()
    except Exception as e:
        logger.warning(f"Failed to record download: {e}")


def get_ocr_results(job_id: str) -> Optional[Dict]:
    """Load OCR results from JSON file"""
    json_path = Config.PROCESSED_DIR / f"{job_id}_ocr.json"
    if not json_path.exists():
        return None

    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================
# TXT Export
# ============================================================

@router.get("/{job_id}/txt")
async def export_txt(job_id: str, include_page_numbers: bool = True, user_id: str = "", request: Request = None):
    """Export OCR results as plain text file"""
    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    lines = []

    for page in ocr_data.get('pages', []):
        page_num = page.get('page_number', 0)

        if include_page_numbers:
            lines.append(f"\n{'='*60}")
            lines.append(f"[Page {page_num}]")
            lines.append(f"{'='*60}\n")

        # Sort by reading order if available
        page_lines = page.get('lines', [])
        if page_lines and page_lines[0].get('reading_order') is not None:
            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))

        for line in page_lines:
            text = line.get('text', '').strip()
            if text:
                lines.append(text)

    content = '\n'.join(lines)

    # Save to file
    output_path = Config.PROCESSED_DIR / f"{job_id}.txt"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)

    if user_id:
        record_download(job_id, user_id, "txt", request.client.host if request and request.client else None)
    return FileResponse(
        path=str(output_path),
        filename=f"{job_id}.txt",
        media_type="text/plain; charset=utf-8"
    )


# ============================================================
# XML Export (ABBYY-compatible format)
# ============================================================

def create_abbyy_xml(ocr_data: Dict, job_id: str) -> str:
    """Create ABBYY FineReader compatible XML"""

    # Root document element
    doc = Element('document')
    doc.set('xmlns', 'http://www.abbyy.com/FineReader_xml/FineReader10-schema-v1.xml')
    doc.set('version', '1.0')
    doc.set('producer', 'BabelBrain OCR')
    doc.set('languages', 'Korean,English')

    # Document data section
    doc_data = SubElement(doc, 'documentData')

    # Paragraph styles
    para_styles = SubElement(doc_data, 'paragraphStyles')
    default_style = SubElement(para_styles, 'paragraphStyle')
    default_style.set('id', '{DEFAULT-STYLE}')
    default_style.set('name', 'Default')
    default_style.set('role', 'text')
    default_style.set('align', 'Left')

    # Process each page
    for page in ocr_data.get('pages', []):
        page_num = page.get('page_number', 1)
        width = page.get('width', 2480)
        height = page.get('height', 3508)

        # Page element
        page_elem = SubElement(doc, 'page')
        page_elem.set('width', str(width))
        page_elem.set('height', str(height))
        page_elem.set('resolution', '400')
        page_elem.set('originalCoords', '1')

        # Text block
        block = SubElement(page_elem, 'block')
        block.set('blockType', 'Text')
        block.set('l', '0')
        block.set('t', '0')
        block.set('r', str(width))
        block.set('b', str(height))

        # Region
        region = SubElement(block, 'region')
        rect = SubElement(region, 'rect')
        rect.set('l', '0')
        rect.set('t', '0')
        rect.set('r', str(width))
        rect.set('b', str(height))

        # Text content
        text_elem = SubElement(block, 'text')
        text_elem.set('id', f'{{PAGE-{page_num}}}')

        # Sort lines by reading order
        page_lines = page.get('lines', [])
        if page_lines and page_lines[0].get('reading_order') is not None:
            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))

        # Group lines into paragraphs (simple grouping by vertical proximity)
        current_par = None
        last_bottom = 0

        for line_data in page_lines:
            bbox = line_data.get('bbox', [0, 0, 100, 20])
            if len(bbox) < 4:
                continue

            l, t, r, b = [int(x) for x in bbox[:4]]
            text = line_data.get('text', '')
            confidence = line_data.get('confidence', 0.95)

            # Start new paragraph if gap is large
            if current_par is None or (t - last_bottom) > 50:
                current_par = SubElement(text_elem, 'par')
                current_par.set('style', '{DEFAULT-STYLE}')

            # Line element
            line_elem = SubElement(current_par, 'line')
            line_elem.set('baseline', str(b))
            line_elem.set('l', str(l))
            line_elem.set('t', str(t))
            line_elem.set('r', str(r))
            line_elem.set('b', str(b))

            # Formatting element
            formatting = SubElement(line_elem, 'formatting')
            formatting.set('lang', 'Korean')
            formatting.set('ff', 'Batang')  # Font family

            # Estimate font size from bbox height
            font_size = max(8, min(72, int((b - t) * 0.7)))
            formatting.set('fs', f'{font_size}.')

            # Character parameters for each character
            if text:
                char_width = (r - l) / max(1, len(text))

                # 실제 CTC 문자별 confidence 사용 (있으면)
                real_char_confidences = line_data.get('char_confidences', [])
                has_real_confidences = bool(real_char_confidences) and len(real_char_confidences) == len(text)

                # 실제 confidence가 없으면 휴리스틱 사용
                if not has_real_confidences:
                    bbox_height = b - t
                    context = {
                        'document_language': 'ko',
                        'bbox_height': bbox_height,
                    }
                    heuristic_confs = char_confidence_estimator.estimate_char_confidences(
                        text, confidence, context
                    )

                for i, char in enumerate(text):
                    char_params = SubElement(formatting, 'charParams')
                    char_l = int(l + i * char_width)
                    char_r = int(l + (i + 1) * char_width)

                    char_params.set('l', str(char_l))
                    char_params.set('t', str(t))
                    char_params.set('r', str(char_r))
                    char_params.set('b', str(b))

                    # 실제 CTC confidence 사용 (0-100 scale)
                    if has_real_confidences:
                        char_conf = int(real_char_confidences[i] * 100)
                        is_suspicious = char_conf < 80
                    elif i < len(heuristic_confs):
                        char_conf_data = heuristic_confs[i]
                        char_conf = int(char_conf_data['confidence'] * 100)
                        is_suspicious = char_conf_data.get('suspicious', False)
                    else:
                        char_conf = int(confidence * 100) if confidence else 95
                        is_suspicious = char_conf < 80

                    char_params.set('charConfidence', str(char_conf))

                    # Mark as suspicious if low confidence
                    if is_suspicious or char_conf < 80:
                        char_params.set('suspicious', '1')

                    # Word boundaries
                    if i == 0 or (i > 0 and text[i-1] == ' '):
                        char_params.set('wordFirst', '1')
                        char_params.set('wordLeftMost', '1')

                    char_params.text = char

            last_bottom = b

    # Pretty print
    xml_str = tostring(doc, encoding='unicode')
    dom = minidom.parseString(xml_str)
    return dom.toprettyxml(indent='  ', encoding=None)


@router.get("/{job_id}/xml")
async def export_xml(job_id: str, user_id: str = "", request: Request = None):
    """Export OCR results as ABBYY-compatible XML file"""
    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    xml_content = create_abbyy_xml(ocr_data, job_id)

    # Save to file
    output_path = Config.PROCESSED_DIR / f"{job_id}.xml"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(xml_content)

    if user_id:
        record_download(job_id, user_id, "xml", request.client.host if request and request.client else None)
    return FileResponse(
        path=str(output_path),
        filename=f"{job_id}.xml",
        media_type="application/xml; charset=utf-8"
    )


# ============================================================
# Excel Statistics Export
# ============================================================

@router.get("/{job_id}/statistics")
async def export_statistics(job_id: str):
    """Export recognition statistics as JSON (for Excel)"""
    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    # Calculate statistics
    stats = {
        'job_id': job_id,
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'total_pages': len(ocr_data.get('pages', [])),
            'total_lines': 0,
            'total_characters': 0,
            'average_confidence': 0,
            'low_confidence_count': 0,  # < 80%
            'suspicious_count': 0,       # < 60%
        },
        'pages': []
    }

    all_confidences = []

    for page in ocr_data.get('pages', []):
        page_stats = {
            'page_number': page.get('page_number', 0),
            'width': page.get('width', 0),
            'height': page.get('height', 0),
            'line_count': len(page.get('lines', [])),
            'character_count': 0,
            'average_confidence': 0,
            'low_confidence_count': 0,
            'suspicious_count': 0,
            'is_multi_column': page.get('is_multi_column', False),
            'confidence_distribution': {
                '90-100': 0,
                '80-90': 0,
                '70-80': 0,
                '60-70': 0,
                '0-60': 0
            }
        }

        page_confidences = []

        for line in page.get('lines', []):
            text = line.get('text', '')
            conf = line.get('confidence', 0.95)

            page_stats['character_count'] += len(text)
            stats['summary']['total_characters'] += len(text)

            if conf is not None:
                conf_pct = conf * 100 if conf <= 1 else conf
                page_confidences.append(conf_pct)
                all_confidences.append(conf_pct)

                # Distribution
                if conf_pct >= 90:
                    page_stats['confidence_distribution']['90-100'] += 1
                elif conf_pct >= 80:
                    page_stats['confidence_distribution']['80-90'] += 1
                elif conf_pct >= 70:
                    page_stats['confidence_distribution']['70-80'] += 1
                elif conf_pct >= 60:
                    page_stats['confidence_distribution']['60-70'] += 1
                else:
                    page_stats['confidence_distribution']['0-60'] += 1

                if conf_pct < 80:
                    page_stats['low_confidence_count'] += 1
                    stats['summary']['low_confidence_count'] += 1

                if conf_pct < 60:
                    page_stats['suspicious_count'] += 1
                    stats['summary']['suspicious_count'] += 1

        stats['summary']['total_lines'] += page_stats['line_count']

        if page_confidences:
            page_stats['average_confidence'] = round(sum(page_confidences) / len(page_confidences), 2)
            page_stats['min_confidence'] = round(min(page_confidences), 2)
            page_stats['max_confidence'] = round(max(page_confidences), 2)

        stats['pages'].append(page_stats)

    if all_confidences:
        stats['summary']['average_confidence'] = round(sum(all_confidences) / len(all_confidences), 2)
        stats['summary']['min_confidence'] = round(min(all_confidences), 2)
        stats['summary']['max_confidence'] = round(max(all_confidences), 2)

    # Recognition rate (percentage of high-confidence text)
    total_lines = stats['summary']['total_lines']
    if total_lines > 0:
        high_conf = total_lines - stats['summary']['low_confidence_count']
        stats['summary']['recognition_rate'] = round(high_conf / total_lines * 100, 2)
    else:
        stats['summary']['recognition_rate'] = 0

    return stats


@router.get("/{job_id}/excel")
async def export_excel(job_id: str, user_id: str = "", request: Request = None):
    """Export recognition statistics as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    # Create workbook
    wb = openpyxl.Workbook()

    # ============ Summary Sheet ============
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary data
    summary_data = [
        ["OCR Recognition Statistics Report"],
        [""],
        ["Job ID", job_id],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [""],
        ["Summary"],
        ["Total Pages", len(ocr_data.get('pages', []))],
        ["Total Lines", sum(len(p.get('lines', [])) for p in ocr_data.get('pages', []))],
        ["Total Characters", sum(sum(len(l.get('text', '')) for l in p.get('lines', [])) for p in ocr_data.get('pages', []))],
    ]

    # Calculate overall confidence
    all_confs = []
    for page in ocr_data.get('pages', []):
        for line in page.get('lines', []):
            conf = line.get('confidence')
            if conf is not None:
                all_confs.append(conf * 100 if conf <= 1 else conf)

    if all_confs:
        avg_conf = sum(all_confs) / len(all_confs)
        low_conf_count = sum(1 for c in all_confs if c < 80)
        recognition_rate = (len(all_confs) - low_conf_count) / len(all_confs) * 100

        summary_data.extend([
            ["Average Confidence", f"{avg_conf:.2f}%"],
            ["Recognition Rate", f"{recognition_rate:.2f}%"],
            ["Low Confidence Lines (< 80%)", low_conf_count],
        ])

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 6:
                cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40

    # ============ Page Details Sheet ============
    ws_pages = wb.create_sheet("Page Details")

    headers = ["Page", "Lines", "Characters", "Avg Confidence", "Min Conf", "Max Conf", "Low Conf Count", "Multi-Column"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_pages.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, page in enumerate(ocr_data.get('pages', []), 2):
        lines = page.get('lines', [])
        confs = [l.get('confidence', 0.95) for l in lines if l.get('confidence') is not None]
        confs = [c * 100 if c <= 1 else c for c in confs]

        char_count = sum(len(l.get('text', '')) for l in lines)

        row_data = [
            page.get('page_number', row_idx - 1),
            len(lines),
            char_count,
            f"{sum(confs)/len(confs):.2f}%" if confs else "N/A",
            f"{min(confs):.2f}%" if confs else "N/A",
            f"{max(confs):.2f}%" if confs else "N/A",
            sum(1 for c in confs if c < 80),
            "Yes" if page.get('is_multi_column') else "No"
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_pages.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws_pages.column_dimensions[get_column_letter(col_idx)].width = 15

    # ============ Confidence Distribution Sheet ============
    ws_dist = wb.create_sheet("Confidence Distribution")

    dist_headers = ["Page", "90-100%", "80-90%", "70-80%", "60-70%", "< 60%"]
    for col_idx, header in enumerate(dist_headers, 1):
        cell = ws_dist.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, page in enumerate(ocr_data.get('pages', []), 2):
        lines = page.get('lines', [])

        dist = {'90-100': 0, '80-90': 0, '70-80': 0, '60-70': 0, '<60': 0}
        for line in lines:
            conf = line.get('confidence')
            if conf is not None:
                conf_pct = conf * 100 if conf <= 1 else conf
                if conf_pct >= 90:
                    dist['90-100'] += 1
                elif conf_pct >= 80:
                    dist['80-90'] += 1
                elif conf_pct >= 70:
                    dist['70-80'] += 1
                elif conf_pct >= 60:
                    dist['60-70'] += 1
                else:
                    dist['<60'] += 1

        row_data = [
            page.get('page_number', row_idx - 1),
            dist['90-100'],
            dist['80-90'],
            dist['70-80'],
            dist['60-70'],
            dist['<60']
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_dist.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            # Color code low confidence
            if col_idx == 6 and value > 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Save
    output_path = Config.PROCESSED_DIR / f"{job_id}_statistics.xlsx"
    wb.save(str(output_path))

    if user_id:
        record_download(job_id, user_id, "excel", request.client.host if request and request.client else None)
    return FileResponse(
        path=str(output_path),
        filename=f"{job_id}_statistics.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ============================================================
# Multi-Format Export (Selective formats)
# ============================================================

class MultiFormatRequest(BaseModel):
    formats: List[str]  # ['pdf', 'txt', 'xml', 'json', 'excel']
    as_zip: bool = True


@router.post("/{job_id}/multi")
async def export_multi_format(job_id: str, request: MultiFormatRequest, user_id: str = "", http_request: Request = None):
    """Export selected formats as ZIP file"""
    import zipfile
    import io

    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    # Validate formats
    valid_formats = {'pdf', 'txt', 'xml', 'json', 'excel'}
    formats = [f.lower() for f in request.formats if f.lower() in valid_formats]
    if not formats:
        raise HTTPException(status_code=400, detail="No valid formats specified")

    # Get original filename
    job_info = job_manager.get_job(job_id)
    base_name = job_id
    if job_info and job_info.get('original_filename'):
        base_name = Path(job_info['original_filename']).stem

    # Create ZIP in memory
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fmt in formats:
            try:
                if fmt == 'pdf':
                    pdf_path = Config.PROCESSED_DIR / f"{job_id}.pdf"
                    if pdf_path.exists():
                        zf.write(pdf_path, f"{base_name}.pdf")

                elif fmt == 'txt':
                    txt_lines = []
                    for page in ocr_data.get('pages', []):
                        page_lines = page.get('lines', [])
                        if page_lines and page_lines[0].get('reading_order') is not None:
                            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))
                        for line in page_lines:
                            text = line.get('text', '').strip()
                            if text:
                                txt_lines.append(text)
                    zf.writestr(f"{base_name}.txt", '\n'.join(txt_lines))

                elif fmt == 'xml':
                    xml_content = create_abbyy_xml(ocr_data, job_id)
                    xml_content = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_content
                    zf.writestr(f"{base_name}.xml", xml_content)

                elif fmt == 'json':
                    json_path = Config.PROCESSED_DIR / f"{job_id}_ocr.json"
                    if json_path.exists():
                        zf.write(json_path, f"{base_name}.json")

                elif fmt == 'excel':
                    # Generate Excel file
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.utils import get_column_letter

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "OCR Statistics"

                        # Add summary data
                        ws['A1'] = "Page"
                        ws['B1'] = "Lines"
                        ws['C1'] = "Characters"
                        ws['D1'] = "Avg Confidence"

                        for row_idx, page in enumerate(ocr_data.get('pages', []), 2):
                            lines = page.get('lines', [])
                            confs = [l.get('confidence', 0.95) for l in lines if l.get('confidence') is not None]
                            confs = [c * 100 if c <= 1 else c for c in confs]

                            ws.cell(row=row_idx, column=1, value=page.get('page_number', row_idx - 1))
                            ws.cell(row=row_idx, column=2, value=len(lines))
                            ws.cell(row=row_idx, column=3, value=sum(len(l.get('text', '')) for l in lines))
                            ws.cell(row=row_idx, column=4, value=f"{sum(confs)/len(confs):.2f}%" if confs else "N/A")

                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_buffer.seek(0)
                        zf.writestr(f"{base_name}_statistics.xlsx", excel_buffer.read())

                    except ImportError:
                        logger.warning("openpyxl not installed, skipping Excel export")

            except Exception as e:
                logger.warning(f"Failed to add {fmt} to export: {e}")

    zip_buffer.seek(0)

    if user_id:
        record_download(job_id, user_id, "zip", http_request.client.host if http_request and http_request.client else None)

    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{base_name}_export.zip"'
        }
    )


# ============================================================
# Batch Export (ZIP with all formats)
# ============================================================

@router.get("/{job_id}/all")
async def export_all(job_id: str):
    """Export all formats as ZIP file"""
    import zipfile
    import tempfile

    ocr_data = get_ocr_results(job_id)
    if not ocr_data:
        raise HTTPException(status_code=404, detail="OCR results not found")

    # Create temp directory for files
    output_dir = Config.PROCESSED_DIR / f"{job_id}_export"
    output_dir.mkdir(exist_ok=True)

    # Get original filename from job
    job_info = job_manager.get_job(job_id)
    base_name = job_id
    if job_info and job_info.get('original_filename'):
        base_name = Path(job_info['original_filename']).stem

    # Generate TXT
    txt_lines = []
    for page in ocr_data.get('pages', []):
        page_lines = page.get('lines', [])
        if page_lines and page_lines[0].get('reading_order') is not None:
            page_lines = sorted(page_lines, key=lambda x: x.get('reading_order', 0))
        for line in page_lines:
            text = line.get('text', '').strip()
            if text:
                txt_lines.append(text)

    txt_path = output_dir / f"{base_name}.txt"
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(txt_lines))

    # Generate XML
    xml_content = create_abbyy_xml(ocr_data, job_id)
    xml_content = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_content

    xml_path = output_dir / f"{base_name}.xml"
    with open(xml_path, 'w', encoding='utf-8') as f:
        f.write(xml_content)

    # Copy PDF if exists
    pdf_path = Config.PROCESSED_DIR / f"{job_id}.pdf"
    if pdf_path.exists():
        import shutil
        shutil.copy(pdf_path, output_dir / f"{base_name}.pdf")

    # Create ZIP
    zip_path = Config.PROCESSED_DIR / f"{job_id}_export.zip"
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_path in output_dir.iterdir():
            zipf.write(file_path, file_path.name)

    # Cleanup temp directory
    import shutil
    shutil.rmtree(output_dir)

    return FileResponse(
        path=str(zip_path),
        filename=f"{base_name}_OCR_results.zip",
        media_type="application/zip"
    )
